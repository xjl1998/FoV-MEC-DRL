from itertools import product as product
from itertools import  combinations
import math
import matlab
import numpy as np
import re
import os
import platform
import random
import matplotlib.pyplot as plt
# from math import log10, log2, cos, sin,exp
import  matlab.engine
import openpyxl
from numpy.ma import log10, log2, cos, sin, exp
from os.path import dirname, abspath, exists

from tqdm import tqdm
import numba as nb
from numba.pycc import CC
cc = CC('yin')
from config import  FLAGS
# FOV的分辨率
#2d的FOV数据大小
FOV_2D = 2*FLAGS.fOV_2DShape[0]*FLAGS.fOV_2DShape[1]
#3d的FOV数据大小
FOV_3D = (4/3)*FOV_2D


#CPU的处理频率
# F_VR = 3 * 10**9
# F_MEC = 10 * 10**9
#
# f_VR = 15
# f_MEC = 15
#
# k_m = 10**(-9)
# k_v = 10**(-9)
#
# E_MEC = 10**(20)
# E_VR = 10**(15)

# np.random.seed(1)
# BW = 40
N_0_dbm = -174 + 10 * log10(FLAGS.BW)
N_0 = np.power(10,(N_0_dbm - 30) / 10)
# N_0 = 10 ** ((N_0_dbm - 30) / 10)
# N_0 =0.00001
# ue_bs_a = 3
# ue_irs_a = 2.2
# irs_bs_a = 2.2
# ue_bs_a = 3.5
# ue_irs_a = 2.5
# irs_bs_a = 2.5
# gfu_bs_a = 3.5#2.5


def G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect):
    '''
    计算综合信道增益G
    h_cue_bs:用户到基站的信道增益
    h_irs_bs：IRS到基站的信道增益，是个一行K列的矩阵
    h_cue_irs：用户到IRS的信道增益，是一个K行一列的矩阵
    reflect：反射矩阵，是一个K行K列的矩阵
    :return:一个综合信道增益的值
    '''
    # print("h_irs_bs",h_irs_bs)
    # print("reflect",reflect)
    h_irs_bs = np.array(h_irs_bs)
    h_cue_irs = np.array(h_cue_irs)
    # temp = np.dot(h_irs_bs.T.conjugate(), reflect)
    temp = h_irs_bs.T.conjugate()
    # print("temp",temp)
    h_cue_irs_bs = np.dot(temp, h_cue_irs)
    G = h_cue_bs + h_cue_irs_bs
    return G


'''

'''
def gen_mec_store_limit( pbd, bs_num,
                         computing_resources,max_power):
        limit = np.zeros_like(computing_resources)
        #计算一个FOV占用的功率
        for fov in range(bs_num):
            p = pbd * 1 * computing_resources[fov]
            limit[fov]=int(max_power)/int(p)
        return limit





'''
每次优先往空余最大的位置放置
'''
def gen_epsilon_rule_larggest(epsilon,storage_limit_mec):
    # bs_num = epsilon.shape[0]
    fov_num = epsilon.shape[1]
    new_epsilon = np.zeros_like(epsilon)

    for fov in range(fov_num):
        selected_bs = epsilon[:,fov]
        storage_limit_mec_copy = storage_limit_mec.copy()
        for time in range(np.sum(selected_bs)):
            max_index = np.argmax(storage_limit_mec_copy)
            storage_limit_mec[max_index]-=1
            storage_limit_mec_copy[max_index] = -1
            if (storage_limit_mec[max_index] == 0):
                storage_limit_mec[max_index]=-1
            new_epsilon[max_index,fov]=1
    return  new_epsilon
    # print("gen_epsilon_rule_larggest")
'''
每次优先往空余最大的位置放置 V2
'''
def gen_epsilon_rule_larggest2(epsilon,storage_limit_mec,BW, G2, omegas, N_0):
    bs_num = epsilon.shape[0]
    fov_num = epsilon.shape[1]
    new_epsilon = np.zeros_like(epsilon)
    storage_limit_mec_back = storage_limit_mec.copy()
    storage_limit_mec_copy = []
    for fov in range(fov_num):
        storage_limit_mec_copy = storage_limit_mec.copy()
        max_index = np.argmax(storage_limit_mec_copy)
        storage_limit_mec[max_index]-=1
        storage_limit_mec_copy[max_index] = -1
        if (storage_limit_mec[max_index] == 0):
         storage_limit_mec[max_index]=-1
        new_epsilon[max_index,fov]=1

    unqualified_fov = []
    for fov in range(fov_num):
        rate = cal_transmit_rate(BW,G2, omegas, fov,new_epsilon,N_0)
        if (rate < 20):
            unqualified_fov.append(fov)
    # fov_con_num = np.ones([fov_num])
    # new_epsilon = np.zeros_like(epsilon)
    detect_time =  0
    while (len(unqualified_fov) > 0 and detect_time<100):
        detect_time+=1
        for fov in unqualified_fov:
            storage_limit_mec_copy = storage_limit_mec.copy()
            for bs in range(bs_num):
                if(new_epsilon[bs,fov]==1):
                    storage_limit_mec_copy[bs]=0
            if (np.sum(storage_limit_mec_copy) == 0):
                return gen_epsilon_rule_larggest(epsilon, storage_limit_mec_back)
            max_index = np.argmax(storage_limit_mec_copy)
            storage_limit_mec[max_index] -= 1
            storage_limit_mec_copy[max_index] = -1
            if (storage_limit_mec[max_index] == 0):
                    storage_limit_mec[max_index] = -1
            new_epsilon[max_index, fov] = 1
        unqualified_fov.clear()
        for fov in range(fov_num):
            rate = cal_transmit_rate(BW, G2, omegas, fov, new_epsilon, N_0)
            if (rate < 20):
                unqualified_fov.append(fov)
    return  new_epsilon
    # print("gen_epsilon_rule_larggest")

from numba import jit
'''
每次优先往空余最小的位置放置 V2
'''
def gen_epsilon_rule_smallest2(epsilon,storage_limit_mec,BW, G2, omegas, N_0):
    bs_num = epsilon.shape[0]
    fov_num = epsilon.shape[1]
    new_epsilon = np.zeros_like(epsilon)
    storage_limit_mec_back = storage_limit_mec.copy()
    storage_limit_mec_copy = []
    for fov in range(fov_num):
        selected_bs = epsilon[:, fov]
        storage_limit_mec_copy = storage_limit_mec.copy()
        for time in range(np.sum(selected_bs)):
            min_index = np.argmin(storage_limit_mec_copy)
            storage_limit_mec[min_index] -= 1
            storage_limit_mec_copy[min_index] = 99
            if (storage_limit_mec[min_index] == 0):
                storage_limit_mec[min_index] = 99
            new_epsilon[min_index, fov] = 1

    unqualified_fov = []
    for fov in range(fov_num):
        rate = cal_transmit_rate(BW,G2, omegas, fov,new_epsilon,N_0)
        if (rate < 20):
            unqualified_fov.append(fov)
    detect_time = 0
    # fov_con_num = np.ones([fov_num])
    # new_epsilon = np.zeros_like(epsilon)
    while (len(unqualified_fov) > 0 and detect_time<100):
        detect_time +=1
        for fov in unqualified_fov:
            storage_limit_mec_copy = storage_limit_mec.copy()
            for bs in range(bs_num):
                if(new_epsilon[bs,fov]==1):
                    storage_limit_mec_copy[bs]=1000
            if (np.sum(storage_limit_mec_copy) == 3000):
                return gen_epsilon_rule_smallest(epsilon, storage_limit_mec_back)
            min_index = np.argmin(storage_limit_mec_copy)
            storage_limit_mec[min_index] -= 1
            storage_limit_mec_copy[min_index] = 99
            if (storage_limit_mec[min_index] == 0):
                storage_limit_mec[min_index] = 99
            new_epsilon[min_index, fov] = 1
        unqualified_fov.clear()
        for fov in range(fov_num):
            rate = cal_transmit_rate(BW, G2, omegas, fov, new_epsilon, N_0)
            if (rate < 20):
                unqualified_fov.append(fov)
    return  new_epsilon


from joblib import Parallel, delayed
'''
穷举法找到当前最优
'''
@cc.export('gen_epsilon_rule_exhaustion', nb.int32(nb.int32[:]))
def gen_epsilon_rule_exhaustion(epsilon,storage_limit_mec,BW, G2, omegas, N_0, fov_sizes,Kb,Ub,ub,cr,total_computing_resources,mec_p_max):
    bs_num = epsilon.shape[0]
    fov_num = epsilon.shape[1]
    new_epsilon = np.zeros_like(epsilon)
    upperlimit = np.sum(epsilon)
    storage_limit_mec_back = storage_limit_mec.copy()
    action_table = gen_action_table_v3(bs_num,fov_num)
    action_space = len(action_table[0])
    min_power = 65536
    min_S  = 65536
    res_r = 0
    avaliable_epsilons = []
    if(fov_num==6):
        for a1 in range(action_space):
            for a2 in range(action_space):
                for a3 in range(action_space):
                    for a4 in range(action_space):
                        for a5 in range(action_space):
                            for a6 in range(action_space):
                                temp_epsilon =  gen_epsilon(bs_num, fov_num,[a1,a2,a3,a4,a5,a6] , action_table)
                                temp_r = 0
                                if(np.sum(temp_epsilon)<=upperlimit):
                                    valid_mark = True
                                    for fov in range(fov_num):
                                        rate = cal_transmit_rate(BW, G2, omegas, fov, temp_epsilon, N_0)
                                        if (rate < 15):
                                            valid_mark=False
                                            break
                                        # temp_r +=(rate-20)
                                    if(valid_mark==True):
                                        total_powers =   cal_total_power(static_power=10, pbd=5e-3, epsilon=temp_epsilon,
                                                            omega=omegas, bs_num=bs_num,
                                                            fov_num=fov_num,
                                                            fov_sizes=fov_sizes, Kb=Kb, Ub=Ub,
                                                            ub=ub, cr=cr,
                                                            computing_resources=total_computing_resources)
                                        # sum_total_powers = np.sum(total_powers)
                                        # mean_total_powers = np.mean(total_powers)
                                        # var_total_powers =  np.var(total_powers)
                                        valid_mark = True
                                        for bs in range(bs_num):
                                            if (total_powers[bs] > mec_p_max):
                                                valid_mark=False
                                                break
                                        # if(min_power>sum_total_powers ):
                                            # min_power = sum_total_powers
                                        if(valid_mark):
                                            avaliable_epsilons.append(temp_epsilon)
                                            # new_epsilon =  temp_epsilon
                                            #     # min_S =var_total_powers
                                            # return new_epsilon
                                        # elif(min_power  == sum_total_powers and min_S>var_total_powers):
                                        #     new_epsilon = temp_epsilon
                                        #     min_S =var_total_powers
                                    # if (np.sum(new_epsilon) == fov_num):
                                    #     return new_epsilon

    elif(fov_num==8):
        for a1 in range(action_space):
            for a2 in range(action_space):
                for a3 in range(action_space):
                    for a4 in range(action_space):
                        for a5 in range(action_space):
                            for a6 in range(action_space):
                                for a7 in range(action_space):
                                    for a8 in range(action_space):
                                        # print([a1,a2,a3,a4,a5,a6,a7,a8])
                                        temp_epsilon = gen_epsilon(bs_num, fov_num, [a1, a2, a3, a4, a5, a6,a7,a8], action_table)
                                        if (np.sum(temp_epsilon) <= upperlimit):
                                            valid_mark = True
                                            for fov in range(fov_num):
                                                rate = cal_transmit_rate(BW, G2, omegas, fov, temp_epsilon, N_0)
                                                if (rate < 15):
                                                    valid_mark=False
                                                    break
                                            if(valid_mark==True):
                                                total_powers =cal_total_power(static_power=10, pbd=5e-3, epsilon=temp_epsilon,
                                                                               omega=omegas, bs_num=bs_num,
                                                                               fov_num=fov_num,
                                                                               fov_sizes=fov_sizes, Kb=Kb, Ub=Ub,
                                                                               ub=ub, cr=cr,
                                                                               computing_resources=total_computing_resources)
                                                for bs in range(bs_num):
                                                    if (total_powers[bs] > mec_p_max):
                                                        valid_mark = False
                                                        break
                                                # if(min_power>sum_total_powers ):
                                                # min_power = sum_total_powers
                                                if (valid_mark):
                                                    # print("找到")
                                                    avaliable_epsilons.append(temp_epsilon)
                                                    # new_epsilon = temp_epsilon
                                                    # min_S =var_total_powers
                                                    # return new_epsilon
                                                # sum_total_powers = np.sum(total_powers)
                                                # mean_total_powers = np.mean(total_powers)
                                                # var_total_powers = np.var(total_powers)
                                                # print(total_powers,var_total_powers)
                                                # if (min_power > sum_total_powers and min_S > var_total_powers):
                                                    # min_power = sum_total_powers
                                                    # new_epsilon = temp_epsilon
                                                    # min_S = var_total_powers

                                                # elif (min_power == sum_total_powers and min_S > var_total_powers):
                                                #     new_epsilon = temp_epsilon
                                                #     min_S = var_total_powers
    return  avaliable_epsilons

'''
每次优先往空余最小的位置放置
'''
def gen_epsilon_rule_smallest(epsilon,storage_limit_mec):
    # bs_num = epsilon.shape[0]
    fov_num = epsilon.shape[1]
    new_epsilon = np.zeros_like(epsilon)

    for fov in range(fov_num):
        selected_bs = epsilon[:,fov]
        storage_limit_mec_copy = storage_limit_mec.copy()
        for time in range(np.sum(selected_bs)):
            min_index = np.argmin(storage_limit_mec_copy)
            storage_limit_mec[min_index]-=1
            storage_limit_mec_copy[min_index]=99
            if(storage_limit_mec[min_index]==0):
                storage_limit_mec[min_index]=99
            new_epsilon[min_index,fov]=1
    return  new_epsilon
    # print("gen_epsilon_rule_smallest")

def zeroforcing(H,max_power):
    # omega = ZF(H, pow)
    K,M= H.shape
    # [K, M] = size(H);
    H_= H.T.conjugate()
    pre = np.dot(H_,np.linalg.inv(np.dot(H,H_)))# pre = H'*inv(H*H');
    pre_=pre.T.conjugate()
    omega = np.sqrt(max_power/np.trace(np.dot(pre,pre_)))*pre
    # omega = sqrt(pow / trace(pre * pre'))*pre;
    return omega
def gen_omega_ZF(NumberOfFov,NumberOfBS,NumberofAntenna,G,max_power):
    omegas = np.zeros([NumberOfBS, NumberOfFov, NumberofAntenna],dtype=np.complex)
    for bs in range(NumberOfBS):
        H = G[bs,:].reshape([NumberOfFov,NumberofAntenna])
        PZF = zeroforcing(H, max_power)
        omegas[bs,:] = PZF.T.conjugate()
    return omegas



def rechoose_epsilon_noCoMP(epsilon,cue_coord,bs_coord,time):
    rechoose_epsilon = np.zeros_like(epsilon)

    for fov in range (rechoose_epsilon.shape[1]):
        # minDist=9999
        # bestBs = -1
        for bs in range(rechoose_epsilon.shape[0]):
            # dist = np.linalg.norm(cue_coord[time,fov,:]-bs_coord[bs,:])
            # if(dist<minDist):
            #     bestBs=bs
            #     minDist = dist
            if(epsilon[bs,fov]==1):
                rechoose_epsilon[bs,fov]=1
                break
    return rechoose_epsilon



def generate_omega_random(bs_num,ue_num,antenna_num,scale_factor=0.2):
    lamb = 1
    omegas = []
    for bs in range(bs_num):
        for ue in range(ue_num):
            omega = np.zeros(shape=[antenna_num],dtype=np.complex)
            for i in range(antenna_num):
                # theta = 2*random.random()*math.pi
                omega[i] = scale_factor*(np.random.rand()*1+np.random.rand()*1j)
            # omegas.append(np.diag(omega))
            omegas.append(omega)
    return np.array(omegas).reshape([bs_num,ue_num,antenna_num])

def generate_omega_fixed(bs_num,ue_num,antenna_num,scale=0.2):
    omegas = scale*(np.ones([bs_num,ue_num,antenna_num])+np.ones([bs_num,ue_num,antenna_num])*1j)
    return omegas

def h_gain_cal(coord_a, coord_b, a,small_fading_style,irs_m):
    '''
    :param coord_a:用户或者基站坐标
    :param coord_b:用户或者基站坐标
    :param a:路径损耗系数
    :param small_fading_style:小尺度衰落
    irs_m:irs元件个数
    :return:增益
    '''
    if small_fading_style == 'Rayleigh':
        small = np.random.normal(0,1/2,1)+np.random.normal(0,1/2,1)*1j
        ad = np.array(np.array(coord_a) - np.array(coord_b)).reshape(1, 3)
        d = np.linalg.norm(ad)
        if d == 0:
            d = 0.000001
        h = np.sqrt(0.001* d ** (-a)) * small
        return h
    else:
        h=[]
        for i in range(irs_m):
            small = np.sqrt(2/3)*(exp(0) * (cos(np.random.rand()*2*np.pi) + sin(np.random.rand()*2*np.pi) * 1j)) \
                    +np.sqrt(1/3)*(np.random.normal(0,1/2,1)+np.random.normal(0,1/2,1)*1j)
            # small=1
            ad = np.array(np.array(coord_a) - np.array(coord_b)).reshape(1, 3)
            d = np.linalg.norm(ad)
            if d == 0:
                d = 0.000001
            h.append(np.sqrt(0.001* d ** (-a)) * small)
        return h
def all_G_gain_cal_MISO(time,bs_num, ue_num,antenna_num,irs_coord,ue_coord,bs_coord,reflect,irs_units_num):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G = []
    count = 0
    for ue in range(ue_num):
        for bs in range(bs_num):
                for antenna in range(antenna_num):
                    h_cue_bs = h_gain_cal(ue_coord[time,ue,:], bs_coord[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_units_num)
                    h_cue_irs = h_gain_cal(ue_coord[time,ue,:], irs_coord, FLAGS.ue_irs_a, "Racian", irs_units_num)
                    h_irs_bs = h_gain_cal(irs_coord, bs_coord[bs], FLAGS.irs_bs_a, "Racian", irs_units_num)
                    if irs_units_num != 0:
                        G.append( G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0])
                    else:
                        G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0])

    return G




# def all_G_gain_cal_MISO_splitI(time,bs_num, ue_num,antenna_num,irs_coord,ue_coord,bs_coord,reflect,irs_units_num):
#     # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
#     G2 = np.zeros([bs_num, ue_num,antenna_num],dtype=np.complex)
#     G = []
#     for ue in range(ue_num):
#         for bs in range(bs_num):
#                 for antenna in range(antenna_num):
#                     h_cue_bs = h_gain_cal(ue_coord[time,ue,:], bs_coord[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_units_num)
#                     h_cue_irs = h_gain_cal(ue_coord[time,ue,:], irs_coord, FLAGS.ue_irs_a, "Racian", irs_units_num)
#                     h_irs_bs = h_gain_cal(irs_coord, bs_coord[bs], FLAGS.irs_bs_a, "Racian", irs_units_num)
#
#                     if irs_units_num != 0:
#                         G2[bs, ue, antenna] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
#                         G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0].real)
#                         G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0].imag)
#                     else:
#                         G2[bs, ue, antenna] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
#                         G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0].real)
#                         G.append(G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0].imag)
#     return G,G2
def all_G_gain_cal_MISO_splitI(time,bs_num, ue_num,antenna_num,irs_coord,ue_coord,bs_coord,reflect,irs_units_num):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G2 = np.zeros([bs_num, ue_num,antenna_num,1],dtype=np.complex)
    g_ue_ris = np.zeros([bs_num, ue_num,irs_units_num,1],dtype=np.complex)
    g_bs_ris = np.zeros([bs_num,irs_units_num,antenna_num],dtype=np.complex)
    g_bs_ue = np.zeros([bs_num, ue_num,antenna_num,1],dtype=np.complex)
    G = []

    for bs in range(bs_num):
                for antenna in range(antenna_num):
                    h_irs_bs = h_gain_cal(irs_coord, bs_coord[bs], FLAGS.irs_bs_a, "Racian", irs_units_num)
                    g_bs_ris[bs, :, antenna] = h_irs_bs
                for ue in range(ue_num):
                    for antenna in range(antenna_num):
                        h_cue_bs = h_gain_cal(ue_coord[time, ue, :], bs_coord[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_units_num)
                        g_bs_ue[bs, ue, antenna] = h_cue_bs
                    h_cue_irs = h_gain_cal(ue_coord[time, ue, :], irs_coord, FLAGS.ue_irs_a, "Racian", irs_units_num)
                    g_ue_ris[bs, ue, :] = h_cue_irs


    return G,G2,g_ue_ris,g_bs_ris,g_bs_ue
def generate_uefov_table(ue_num):
    '''简化模型，目前认为每个用户请求渲染的fov是一致的'''
    rs = np.zeros(ue_num,dtype=np.int)
    for i in range(ue_num):
        rs[i]=i
    return rs
def generate_bsfov_table(epsilon):
    [bs_num,fov_num]=epsilon.shape
    bsfov_table = []
    for fov in range(fov_num):
        bss = []
        for bs in range(bs_num):
            if(epsilon[bs,fov]==1):
                bss.append(bs)
        bsfov_table.append(bss)
    return bsfov_table



def cal_sinr(G,omegas,current_fov,epsilon,N0):
    sum_up=0
    sum_down=N0
    bs_num = omegas.shape[0]
    ue_num = omegas.shape[1]
    antenna_num = G.shape[2]
    for bs in range(bs_num):
            for fov in range(ue_num):
                if (epsilon[bs,fov] == 1):
                    g = G[bs, current_fov, :].reshape([antenna_num, 1])
                    omega = np.array(omegas[bs, fov, :]).reshape([antenna_num, 1]).T.conjugate()
                    item = np.power(np.linalg.norm(np.dot(omega, g)), 2)
                    if(fov==current_fov):
                        sum_up += item
                    else:
                        sum_down+=item
        # else:
        #     conn = epsilon[:, fov].flatten()
        #     for bs in range(bs_num):
        #         if (conn[bs] == 1):
        #             g = G[bs, fov, :]
        #             omega = np.array(omegas[bs, current_fov, :]).T
        #             sum_down += np.power(np.linalg.norm(np.dot(omega, g)), 2)

    return sum_up/sum_down

def cal_transmit_rate(bw,G,omegas,current_fov,epsilon,N0):
    sinr =cal_sinr(G,omegas,current_fov,epsilon,N0)
    return bw*np.log2((1+sinr))
def cal_total_rendered_fov_sizes(fov_sizes,cr):
    rs = np.zeros(len(fov_sizes))
    for fov in fov_sizes:
        rs[fov]=cal_rendered_size(fov_sizes[fov], cr)
    return  rs
def cal_total_computing_resources(fov_sizes,Kb,Ub,ub,cr):
    computing_resources = np.zeros(len(fov_sizes))
    for fov in range(len(fov_sizes)):
        computing_resources[fov]=cal_rendered_computing_resources(cal_rendered_size(fov_sizes[fov],cr),Kb,Ub,ub)
    return computing_resources
def cal_transmit_power(epsilon,omega,bs_num,fov_num):
    power = np.zeros(bs_num,dtype=np.float)
    for bs in range(bs_num):
        p=0
        for fov in range(fov_num):
            p += epsilon[bs,fov]*np.power(np.linalg.norm(omega[bs,fov,:]),2)
        power[bs]=p
    return power
def cal_rendered_power(pbd,epsilon,bs_num,fov_num,computing_resources):
    power=np.zeros(bs_num)
    for bs in range(bs_num):
        p=0
        for fov in range(fov_num):
            p+=(pbd*epsilon[bs,fov]*computing_resources[fov])
        power[bs]=p
    return power
def cal_total_power(static_power,pbd,epsilon,omega,bs_num,fov_num,fov_sizes,Kb,Ub,ub,cr):

    computing_resources = cal_total_computing_resources(fov_sizes, Kb, Ub, ub,cr)
    transmit_power = cal_transmit_power(epsilon,omega,bs_num,fov_num)
    rendered_power = cal_rendered_power(pbd,epsilon,bs_num,fov_num,computing_resources)
    return transmit_power+rendered_power+np.ones(bs_num)*static_power

def cal_total_power(static_power,pbd,epsilon,omega,bs_num,fov_num,fov_sizes,Kb,Ub,ub,cr,computing_resources):


    transmit_power = cal_transmit_power(epsilon,omega,bs_num,fov_num)
    rendered_power = cal_rendered_power(pbd,epsilon,bs_num,fov_num,computing_resources)
    return transmit_power+rendered_power+np.ones(bs_num)*static_power

def cal_render_power(static_power,pbd,epsilon,bs_num,fov_num,fov_sizes,Kb,Ub,ub,cr,computing_resources):

    rendered_power = cal_rendered_power(pbd,epsilon,bs_num,fov_num,computing_resources)
    return rendered_power+np.ones(bs_num)*static_power

def generate_max_computing_resources_mec(mec_num,min_size,max_size):
    '''
    :param mec_num:
    :param min_size:
    :param max_size:
    :return:
    '''
    rs = np.zeros([mec_num],np.int)
    for i in range(mec_num):
        rs[i] = random.randint(min_size,max_size)
    return rs
def generate_storage_mec(mec_num,min_size,max_size):
    '''
    :param mec_num:
    :param min_size:
    :param max_size:
    :return:
    '''
    rs = np.zeros([mec_num], np.int)
    for i in range(mec_num):
        rs[i] = random.randint(min_size, max_size)
    return rs
def generate_fov_size(fov_num,min_size,max_size):
    '''
    :param fov_num:
    :param min_size:
    :param max_size:
    :return:
    '''

    rs = np.zeros([fov_num],dtype=np.int)
    for i in range(fov_num):
        rs[i] = random.randint(min_size, max_size)
    return rs
def cal_rendered_size(fov_size,cr):
    '''
    :param fov_size:
    :param cr: 压缩系数
    :return:
    '''
    rs = 3*8*np.power(fov_size,2)*2/cr
    return rs
def cal_rendered_computing_resources(rendered_fov_size,Kb,Ub,ub):
    '''
    :param rendered_fov_size:
    :param Kb: the architecture coefficient, which is related to the CPU architecture of MEC
    :param Ub: the operating frequency of the MEC
    :param ub: the number of CPU cycles required by MEC to process data per bit during rendering
    :return:
    '''
    rs = rendered_fov_size*Kb*np.power(Ub,2)*ub
    return rs


'''
计算beamforming中的omega
'''
def omega_cal(h):
    omega = np.zeros_like(h).astype(dtype=np.complex)

    return omega

def calculate_reward(caching_matrix):
    reward = 0
    dic = {}.fromkeys(caching_matrix)
    if(len(dic)==len(caching_matrix)):
        reward = -100
    else:
        reward = -500
    return  reward



def dist_calc_x(user_state,x,dist,angle,angle_fix,x_min,x_max,max_speed):
        '''
        Use for calculating the distance of movement in each TS in terms of X-axis
        :param user_state:
        :param x:
        :param dist:
        :param angle:
        :param angle_fix:
        :param x_min:
        :param x_max:
        :param max_speed:
        :return:
        '''
        if user_state == 0:  # random
            new_x = x + dist * math.cos(angle)
        else:
            new_x = x + dist * math.cos(angle) + 4/5 * max_speed * math.cos(angle_fix)

        while new_x < x_min or new_x > x_max:
            # print('edge_x')
            new_angle = np.pi + angle
            if user_state == 0:  # random
                new_x = x + dist * math.cos(new_angle)
            else:
                new_x = x + dist * math.cos(new_angle) + 4/5 * max_speed * math.cos(angle_fix)
        return new_x

def dist_calc_y(user_state,y,dist,angle,angle_fix,y_min,y_max,max_speed):
        '''
        Use for calculating the distance of movement in each TS in terms of Y-axis
        :param user_state:
        :param x:
        :param dist:
        :param angle:
        :param angle_fix:
        :param x_min:
        :param x_max:
        :param max_speed:
        :return:
        '''
        if user_state == 0:  # random
            new_y = y + dist * math.sin(angle)
        else:
            new_y = y + dist * math.sin(angle) + 4/5 * max_speed * math.sin(angle_fix)
        while new_y< y_min or new_y > y_max:
            # print('edge_y')
            new_angle = - angle
            if user_state == 0:  # random
                new_y = y + dist * math.sin(new_angle)
            else:
                new_y = y + dist * math.sin(new_angle) + 4/5 * max_speed * math.sin(angle_fix)
        return new_y


def all_G_gain_cal(bs_num, irs_coord,ue_num,coord_a,coord_b,reflect,irs_m):
    # channel_space=np.array(channel_space).reshape(cuenum+chnum,chnum)
    G =np.zeros([bs_num,ue_num],dtype=np.complex)
    # G = np.zeros(bs*cuenum, dtype="complex")
    count = 0
    for ue in range(ue_num):
        for bs in range(bs_num):
                h_cue_bs = h_gain_cal(coord_a[ue], coord_b[bs], FLAGS.gfu_bs_a, "Rayleigh", irs_m)
                h_cue_irs = h_gain_cal(coord_a[ue], irs_coord, FLAGS.ue_irs_a, "Racian", irs_m)
                h_irs_bs = h_gain_cal(irs_coord, coord_b[bs], FLAGS.irs_bs_a, "Racian", irs_m)
                # test1=(cuenum_i + 1) * (bs_i + 1) * (chnum_i + 1) - 1
                if irs_m != 0:
                    G[bs,ue] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
                else:
                    G[bs,ue] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
                count +=1
    return G

    # for c in range(cuenum):
    #     if channel_space[c] != 0:
    #         which_bs = int(channel_space[c] / chnum)
    #         which_ch = channel_space[c] % chnum
    #         h_cue_bs = h_gain_cal(coord_a[c], coord_b[which_bs], gfu_bs_a, "Rayleigh", irs_m)
    #         h_cue_irs = h_gain_cal(coord_a[c], irs_coord, ue_irs_a, "Racian", irs_m)
    #         h_irs_bs = h_gain_cal(irs_coord, coord_b[which_bs], irs_bs_a, "Racian", irs_m)
    #         if irs_m != 0:
    #             G[c] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0][0]
    #         else:
    #             G[c] = G_gain_cal(h_cue_bs, h_irs_bs, h_cue_irs, reflect)[0]
    #     else:
    #         G[c] = 0
    # return G
def sic_decode_judge(cuenum, channel_num, G):
    '''
    判断解码顺序是否满足条件
    '''
    for i in range(channel_num):
        for j in range(cuenum):
            if pow(abs(G[cuenum+i][i]),2)>pow(abs(G[j][i]),2) and G[j][i]!=0:
                return -1
    return 1

def r_min_judge(cue, ch_k, G, action_c_ch, action_c_p, r_min):
    for i in range(cue):
        if sum(action_c_p[i,:]) != 1:
            return -1, None
    for i in range(ch_k):
        if sum(action_c_p[:, i]) > 1:
            return -1, None
    r_arr=[]
    for i in range(cue):
        for j in range(ch_k):
            if action_c_ch[i][j] == 1:
                r = FLAGS.BW*log2(1+(pow(abs(G[i][j]),2)*action_c_p[i,j]/(N_0)))
                r_arr.append(r)
    for i in range(cue):
        if r_arr[i] < r_min:
            return -1, None
    return 1, r_arr








    # for i in range(cue):
    #     r_i=0
    #     for j in range(ch_k):
    #         for k in range(ch_k):
    #             if G[cue+k,j]!=0:
    #                 G_gbu_bs=G[cue+k,j]
    #                 temp=cue+k
    #                 break
    #         i_gbu=action_c_p[temp,j]*pow(abs(G_gbu_bs),2)
    #         i_gfu=0
    #         for k in range(cue):
    #             i_gfu=i_gfu+action_x[j][k][i]*pow(abs(G[k,j]),2)*action_c_p[k,j]
    #         r_i_channel=BW*log2(1+(pow(abs(G[i,j]),2)*action_c_p[i,j]/(i_gfu+i_gbu+N_0)))
    #         r_i=r_i+r_i_channel
    #     r_arr.append(r_i)
    # for i in range(ch_k):
    #     for j in range(ch_k):
    #         if i==j:
    #             r_i=BW*log2(1+(pow(abs(G[cue+i,j]),2)*action_c_p[cue+i,j]/(N_0)))
    #             r_arr.append(r_i)
    # for i in range(cue):
    #     if r_arr[i] < r_min:
    #         return -1
    # return 1
def clean_G(G,cuenum,tau,channel_num):
    for i in range(cuenum):
        for j in range(channel_num):
            if pow(abs(G[i,j]),2)<tau:
                G[i,j]=0
    # for j in range(channel_num):
    #     for i in range(cuenum):
    #         if pow(abs(G[i,j]),2)<pow(abs(G[cuenum+j,j]),2):
    #             G[i,j]=0
    return G
def tau_judge_fun(G,channel_num,cuenum,tau):
    for cue in range(cuenum):
        if sum(G[cue,:])==0:
            return -1
    return 1
# def fun(cue, ch_k, r_arr, action_c_p):
#     '''
#     计算r
#     '''
#     t_relay = 0
#     for i in range(cue):
#         for j in range(ch_k):
#             if action_c_p[i][j] != 0:
#                 t_relay += C/(cr*r_arr[i])
#     return t_relay
#     # G_gbu_bs=0
#     # G=np.array(G).reshape(cue+ch_k,ch_k)
#     # action_c_p=np.array(action_c_p).reshape(cue+ch_k,ch_k)
#     # temp=0
#     # r=0
#     # throughput=0
#     # r_arr=[]
#     # for i in range(cue):
#     #     r_i=0
#     #     for j in range(ch_k):
#     #         for k in range(ch_k):
#     #             if G[cue+k,j]!=0:
#     #                 G_gbu_bs=G[cue+k,j]
#     #                 temp=cue+k
#     #                 break
#     #         i_gbu=action_c_p[temp,j]*pow(abs(G_gbu_bs),2)
#     #         i_gfu=0
#     #         for k in range(cue):
#     #             i_gfu=i_gfu+action_x[j][k][i]*pow(abs(G[k,j]),2)*action_c_p[k,j]
#     #         r_i_channel=BW*log2(1+(pow(abs(G[i,j]),2)*action_c_p[i,j]/(i_gfu+i_gbu+N_0)))
#     #         r_i=r_i+r_i_channel
#     #     throughput=throughput+r_i
#     #     r_arr.append(r_i)
#     #     # r_i=r_i-r_min
#     #     r=r+r_i
#     # # print(throughput,r)
#     # # print("fun_r_arr",r_arr)
#     # return throughput,r,r_arr

def render_time(loc):
    time = 0
    if loc == 0:
        time = (FOV_2D*FLAGS.f_vr)/(FLAGS.F_vr)
    else:
        time = (FOV_2D*FLAGS.f_mec)/(FLAGS.F_mec)
    #时间单位换算成毫秒
    return time*10**3



def qos_judge(G, r_min, cuenum, channel_num, action_c_ch, action_c_p):
    r_min_judge_va, r_arr = r_min_judge(cuenum,channel_num,G,action_c_ch,action_c_p,r_min)
    # sic_decode_result =sic_decode_judge(cuenum, channel_num, G)
    # if r_min_judge_va==-1 or tau_judge==-1 or sic_decode_result == -1 :
    if r_min_judge_va== -1 :
        print("r_min_judge_va",r_min_judge_va)
        return -1, None
    return 1, r_arr
def channel_space_generate(ch_k):
    '''
    可用的各种信道矩阵
    '''
    available_space_channel = []
    a = range(2)#a可以看作列表[0,1]
    for item in product(a, repeat=ch_k):# product(range(2), repeat=3) --> 000 001 010 011 100 101 110 111,笛卡尔积product(A,repeat=3)等价于product(A,A,A)，product(A, B) 和 ((x,y) for x in A for y in B)一样.
        available_space_channel.append(item)
    return available_space_channel

def reflect_calculate(reflect_action_arr,reflect_amp_arr,irs_m):
    '''
    :param reflect_action_arr: 反射矩阵角度
    :return: 返回计算后的反射矩阵
    '''
    reflect = np.zeros((irs_m,irs_m),dtype=np.complex)
    for i in range(irs_m):
        reflect[i,i] = reflect_amp_arr[i]*exp(0) * (cos(reflect_action_arr[i]) + sin(reflect_action_arr[i]) * 1j)
    return reflect
def x_generate(G,ch_k,cue):
    '''
    可用的各种反射矩阵
    '''
    G = np.array(G).reshape(cue, ch_k)
    action_x = []
    # action_x=np.zeros((cue+ch_k,cue+ch_k,ch_k))
    for i in range(ch_k):
        action_x_ch = np.zeros((cue, cue+ch_k))
        temp_arr = np.array(G[:, i])
        for j in range(cue):
            if temp_arr[j] == 0:
                action_x_ch[j,:]=0
            for k in range(cue):
                if temp_arr[j] != 0 and pow(abs(temp_arr[j]), 2) <= pow(abs(temp_arr[k]), 2):
                    action_x_ch[j][k]=1
                if temp_arr[j] != 0 and temp_arr[k] == 0:
                    action_x_ch[j][k] = 1
                if j == k:
                    action_x_ch[j][k] = 0
        action_x.append(action_x_ch)
    return action_x
def plot_mode_irs_compare(number,ave_throughput_arr1,throughput_arr1,in_ave_throughput_arr1,stat_lst):
    plt.figure()
    for stat in range(len(stat_lst)):
        in_ave_throughput_arr = in_ave_throughput_arr1[stat]
        if stat==0:
            plt.plot(np.arange(0,len(in_ave_throughput_arr),9), in_ave_throughput_arr[::9],c='#00CED1' ,marker='*',alpha=0.4, label='C=3')#np.arange函数返回一个有终点和起点的固定步长的排列
        if stat==1:
            plt.plot(np.arange(0,len(in_ave_throughput_arr),9), in_ave_throughput_arr[::9],c='#9932CC', marker='<',alpha=0.4, label='C=2')#np.arange函数返回一个有终点和起点的固定步长的排列
        if stat==2:
            plt.plot(np.arange(0,len(in_ave_throughput_arr),9), in_ave_throughput_arr[::9],c='g', marker='>',alpha=0.4, label='C=1')
        if stat==3:
            plt.plot(np.arange(0,len(in_ave_throughput_arr),9), in_ave_throughput_arr[::9],c='#DC143C', marker='o',alpha=0.4, label='NO_IRS')
    plt.grid(linestyle='-.')
    plt.ylabel('interval_ave_Throughput')
    plt.xlabel('Steps')
    dir_path=dirname(abspath(__file__))
    plt.legend(loc = 'best')
    plt.savefig(dir_path + '/convergence.pdf')#, dpi=300)
    plt.show()
    plt.figure()
    for stat in range(len(stat_lst)):
        ave_throughput_arr = ave_throughput_arr1[stat]
        if stat==0:
            plt.plot(np.arange(10000,len(ave_throughput_arr),9999), ave_throughput_arr[10000:210000:9999],c='#00CED1' ,marker='*',alpha=0.4, label='C=3')#np.arange函数返回一个有终点和起点的固定步长的排列
        if stat==1:
            plt.plot(np.arange(10000,len(ave_throughput_arr),9999), ave_throughput_arr[10000:210000:9999],c='#9932CC', marker='<',alpha=0.4, label='C=2')#np.arange函数返回一个有终点和起点的固定步长的排列
        if stat==2:
            plt.plot(np.arange(10000,len(ave_throughput_arr),9999), ave_throughput_arr[10000:210000:9999],c='g', marker='>',alpha=0.4, label='C=1')
        if stat==3:
            plt.plot(np.arange(10000,len(ave_throughput_arr),9999), ave_throughput_arr[10000:210000:9999],c='#DC143C', marker='o',alpha=0.4, label='NO_IRS')
    plt.grid(linestyle='-.')
    plt.ylabel('Ave_Throughput')
    plt.xlabel('Steps')
    dir_path=dirname(abspath(__file__))
    plt.legend(loc = 'best')
    plt.savefig(dir_path + '/convergence1.pdf')#, dpi=300)
    plt.show()
    plt.close()
def npyload(filename):
    """
    :功能：读取npy文件
    :param filename: 文件名称
    :return:
    """
    print('read file: %s' % (filename))
    return np.load(filename, allow_pickle=True).item()
def npysave(data, filename):
    """
    :功能：保存npy文件
    :param data: 数据
    :param filename: 文件名
    :return:
    """
    namearr = re.split(r'[\\/]', filename)

    #   判断操作系统
    sys = platform.system()
    if sys == "Windows":
        pathstr = '\\'.join(namearr[:-1])
    elif sys == "Linux":
        pathstr = '/'.join(namearr[:-1])
    filestr = namearr[-1]
    if not os.path.exists(pathstr):
        print('make dir：%s' % (pathstr))
        os.makedirs(pathstr)
    print('write to: %s' % (filename))
    np.save(filename, data)
def excel_save(excel,irs_m,stat):
    wb=openpyxl.Workbook()
    ws=wb.create_sheet("sheet1")
    i=1
    for a in excel:
        for j in range(irs_m):
            ws.cell(row=i,column=j+1).value=a[0][j]
        i+=1
    wb.save('data'+stat+'.xlsx')
# if __name__ == "__main__":
# channel_space_generate(2,3,2,2)
# p_space_generate(2,3,2,2,2)
#     reflect_space_generate(2,2)
#     sic_order_calculate([1,1,0],3, 2,  [[3,3,0],[5,5,0],[7,7,0],[4,3,0],[2,2,0]], [0,0,0],  [[1,1],[0,1],[1,0],[0,1],[1,0]],  [[6.28318531,0], [0,6.28318531]],2)
#     x_generate([[1,1.1], [2.1,0], [0,0.1], [0,0.1], [1,0]],2,3)
#     fun(3,2,[[1,1.1],
#              [2.1,0],
#              [0,0.1],
#              [0,0.1],
#              [1,0]],
#         [[[1,1],[1,1],[1,0],[1,0],[1,1]],[[0,0],[1,0],[1,0],[1,0],[0,0]],[[0,1],[0,1],[0,1],[0,1],[0,1]],
#                [[0,1],[0,1],[0,1],[0,1],[0,1]],[[1,0],[1,0],[1,0],[1,0],[1,0]]],[[1,1],
#                                                                                  [1,0],
#                                                                                  [0,1],
#                                                                                  [0,1],
#
#                                                                                  [1,0]])
def ch_max_cue_judge(action_c_p,cuenum,ch_k,gfu_max):
    judge=0
    # print("p_max",p_max)
    for i in range(ch_k):
        count = 0
        for j in range(cuenum):
            if action_c_p[j][i]!=0:
                count+=1
        if count>gfu_max:
            judge=1
            break
    if judge ==1:
        return -1
    else:
        return 1
def cue_max_ch_judge(action_c_p,cuenum,ch_k,ch_max):
    judge=0
    # print("p_max",p_max)
    for i in range(cuenum):
        count = 0
        for j in range(ch_k):
            if action_c_p[i][j]!=0:
                count+=1
        if count>ch_max:
            judge=1
            break
    if judge ==1:
        return -1
    else:
        return 1


def average_power(ue,bs, ch, ch_space, p_max):
    power_result = np.zeros(ue)
    bs_power = np.zeros((ue,2))
    for i in range(ue):
        which_bs = int(ch_space[i] / ch)
        bs_power[ue][0] = ue
        bs_power[ue][1] = which_bs
    # for j in range(bs):
    k = 0
    for i in range(ch_space.shape[0]):
        for j in range(ch_space.shape[1]):
            if ch_space[i][j] != 0:
                k += 1
    average_p = p_max/k
    for i in range(ue):
        for j in range(ch):
            if ch_space[i][j] != 0:
                power_result[i][j] = average_p
    return power_result

def gen_epsilon(bs_num,fov_patch_num):
    avg_count=int(fov_patch_num /bs_num)
    rest = fov_patch_num %bs_num
    if(rest==0):
        avg_count = avg_count+rest
    epsilon = np.zeros([bs_num, fov_patch_num])
    begin_index=0
    for bs in range(bs_num):
        for i in range(avg_count):
            epsilon[bs][begin_index]=1
            if(begin_index+1<fov_patch_num):
                begin_index=begin_index+1
    print("初始化了一个符合条件的epsilon")
    return epsilon
def gen_epsilon_V2(bs_num,fov_patch_num):
    avg_count=int(fov_patch_num /bs_num)
    rest = fov_patch_num %bs_num
    if(rest==0):
        avg_count = avg_count+rest
    epsilon = np.zeros([bs_num, fov_patch_num])
    begin_index=0
    for bs in range(bs_num):
        for i in range(avg_count):
            epsilon[bs][begin_index]=1
            if(begin_index+1<fov_patch_num):
                begin_index=begin_index+1
    print("初始化了一个符合条件的epsilon")
    return epsilon
def change_epsilon_by_actions(epsilon,actions,bs_num,fov_patch_num):
    for bs in range(bs_num):
        action = actions[bs]
        if(action == 2*fov_patch_num):
            return  epsilon
        if(action<fov_patch_num):
            add_action = action
            epsilon[bs][add_action]=1
        else:
            sub_action = action-fov_patch_num
            epsilon[bs][sub_action]=0
    return epsilon

def gen_epsilon(bs_num,fov_num,fov_actions,action_tables):
    epsilon = np.zeros([bs_num,fov_num],dtype=np.int)
    for f in range(fov_num):

        for bs in range(bs_num):
            fov_action = action_tables[bs][fov_actions[f]]
            epsilon[bs][f] = fov_action[bs]
    return epsilon
@cc.export('gen_action_table_v2', nb.int32(nb.int32[:]))
def gen_action_table_v2(bs_num,fov_num):
    table = []
    tables=[]
    if(bs_num==2):
        table.append([0, 1])
        table.append([1,0])
        table.append([1, 1])
    if(bs_num==3):
        table.append([0, 0, 1])
        table.append([0, 1, 0])
        table.append([0, 1, 1])
        table.append([1, 0, 0])
        table.append([1, 0, 1])
        table.append([1, 1, 0])
        table.append([1, 1, 1])
    elif(bs_num==4):
        table.append([0, 0,0, 1])
        table.append([0, 0,1, 0])
        table.append([0, 0,1, 1])
        table.append([0,1, 0, 0])
        table.append([0,1, 0, 1])
        table.append([0,1, 1, 0])
        table.append([0,1, 1, 1])
        table.append([1, 0, 0, 0])
        table.append([1, 0, 0, 1])
        table.append([1, 0, 1, 0])
        table.append([1, 0, 1, 1])
        table.append([1, 1,0, 0])
        table.append([1, 1, 0, 1])
        table.append([1, 1, 1, 0])
        table.append([1, 1, 1, 1])

    for  i in range(fov_num):
        tables.append(table)
    return tables


@cc.export('gen_action_table_v3', nb.int32(nb.int32[:]))
def gen_action_table_v3(bs_num,fov_num):
    table = []
    tables=[]
    if(bs_num==2):
        table.append([0, 1])
        table.append([1,0])
        table.append([1, 1])
    if(bs_num==3):
        table.append([0, 0, 1])
        table.append([1, 0, 0])
        table.append([0, 1, 0])
        table.append([1, 1, 0])
        table.append([0, 1, 1])
        table.append([1, 0, 1])
        table.append([1, 1, 1])

    elif(bs_num==4):
        table.append([1, 0, 0, 0])
        table.append([0, 0,0, 1])
        table.append([0, 0,1, 0])
        table.append([0,1, 0, 0])
        table.append([0, 0,1, 1])
        table.append([0,1, 0, 1])
        table.append([0,1, 1, 0])
        table.append([0,1, 1, 1])
        table.append([1, 0, 0, 1])
        table.append([1, 0, 1, 0])
        table.append([1, 0, 1, 1])
        table.append([1, 1,0, 0])
        table.append([1, 1, 0, 1])
        table.append([1, 1, 1, 0])
        table.append([1, 1, 1, 1])

    for  i in range(fov_num):
        tables.append(table)
    return tables

def available_space(bs_num,fov_num):
    available_space=np.zeros(bs_num,dtype=np.int)
    rest= fov_num%bs_num
    avg_num=int(fov_num/bs_num)
    for bs in range(bs_num):
        available_space[bs]=avg_num
    available_space[bs_num-1]=available_space[bs_num-1]+rest
    return  available_space

def gen_action_space(num,min_dim,max_dim):
    rs = []

    a = range(num)  # a可以看作列表[0,1]
    for i in range(min_dim,max_dim+1):
        comb=list (combinations(a,i))
        for set in comb:
            rs.append(set)

    return rs
